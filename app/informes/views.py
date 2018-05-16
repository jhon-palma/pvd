from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from django.contrib import messages
from pvd.validators import *
from app.asistencia.models import *
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.drawing.image import Image
import openpyxl
from django.core.files.storage import FileSystemStorage
import base64
from django.template.loader import render_to_string
from weasyprint import HTML

@login_required(login_url="/")
def subir_soporte(request):
    create = True
    sedes = Sede.objects.all()
    if request.method == 'POST':
        validator = SoporteValidator(request.POST)
        validator.required = ['fecha']
        if validator.is_valid():
            soporte = Soporte()
            soporte.soporte = request.FILES['soporte']
            soporte.fecha = request.POST['fecha']
            soporte.sede = Sede.objects.get(pk=request.POST['sede'])
            ahora = datetime.datetime.now()
            soporte.soporte.name = "%s-%s.%s" % (soporte.fecha,ahora.minute,'pdf')
            soporte.save()
            messages.success(request,validator.getMessage())
            return HttpResponseRedirect('/informes/subir_soporte')
        else:
            messages.warning(request,validator.getMessage())
            return HttpResponseRedirect('/informes/subir_soporte')
    return render(request,'informes/subir_soporte.html',{'create':create,'sedes':sedes})

@login_required(login_url="/")
def consulta_soporte(request):
    query = True
    sedes = Sede.objects.all()
    if request.method == 'POST':
        validator = SoporteValidator(request.POST)
        validator.required = ['fecha']
        if validator.is_valid():
            edit = True
            fecha = request.POST['fecha']
            soportes = Soporte.objects.filter(fecha = fecha, sede =request.POST['sede'])
            return render(request,'subir_soporte.html', {'edit':edit, 'soportes':soportes,'fecha':fecha })
        else:
            messages.warning(request,validator.getMessage())
            return HttpResponseRedirect('/informes/consulta_soporte',{'query':query,'sedes':sedes})
    return render(request,'informes/subir_soporte.html',{'query':query,'sedes':sedes})

@login_required(login_url="/")
def report_sexo(request):
    query = True
    cursos = Curso.objects.all()
    return render(request,'informes/report_sexo.html',{'query':query, 'cursos':cursos})

@login_required(login_url="/")
def get_genero(request,*args,**kwargs):

    fi = request.GET['fechaInicial']
    ff = request.GET['fechaFinal']
    mujeres = Asistencia.objects.filter(fecha__range=[fi,ff],persona__sexo__contains="F")
    hombres = Asistencia.objects.filter(fecha__range=[fi,ff],persona__sexo__contains="M")

    br1 = mujeres.count()
    br2 = hombres.count()

    if br1 > br2:
        limite = round(br1/10.)*10
        inferior = br2
    elif br2 > br1:
        limite = round(br2/10.)*10
        inferior = br1
    else:
        limite = round(br2/10.)*10
        inferior = br1

    total = br1 + br2

    default = [br1, br2 ]
    labels = ['Mujeres','Hombres']

    data = {
        "labels": labels,
        "default": default,
        "total": total,
        "limite":limite,
        "inferior":inferior,
    }
    return JsonResponse(data)

@login_required(login_url="/")
def report_curso(request):
    query = True
    cursos = Curso.objects.all()

    if request.method == 'POST':
        create = True
        fi = request.POST['fechaInicial']
        ff = request.POST['fechaFinal']
        curso = request.POST['curso']
        cursos = Curso.objects.get(pk = request.POST['curso'] )
        asistentes = PersonaCurso.objects.filter(curso=curso).order_by('persona__apellidos')

        asistencia = Asistencia.objects.filter(curso=curso,fecha__range=[fi,ff]).order_by('persona__apellidos')
        fechas = Asistencia.objects.filter(curso=curso,fecha__range=[fi,ff]).values_list('fecha', flat=True).distinct()

        return render(request,'informes/report_curso.html',{'asistentes':asistentes,'fi':fi,'ff':ff,'cursos':cursos,'asistencia':asistencia,'fechas':fechas,'create':create })

    return render(request,'informes/report_curso.html',{'query':query, 'cursos':cursos})

@login_required(login_url="/usuarios/login/")
def report_datos(request):
    if request.method == 'POST':
        fi = request.POST['fechaInicial']
        ff = request.POST['fechaFinal']

        personas = Asistencia.objects.filter(fecha__range=[fi,ff])
        #personas = Asistencia.objects.distinct('persona').all

        wb = Workbook()
        ws = wb.active

        celdFill = PatternFill(start_color='2E64FE', end_color='2E64FE', fill_type='solid')

        if 'estilo' not in wb.named_styles:
            estilo = NamedStyle(name="estilo")
            estilo.font = Font(color='FFFFFF', bold=True)
            wb.add_named_style(estilo)
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=18):
            for cell in rows:
                cell.style = 'estilo'
                cell.fill = celdFill

        ws['A1'] = 'TIPO_IDENTIFICACION'
        ws['B1'] = 'NUMERO_IDENTIFICACION'
        ws['C1'] = 'NOMBRES'
        ws['D1'] = 'APELLIDOS'
        ws['E1'] = 'EMAIL'
        ws['F1'] = 'EMAIL2'
        ws['G1'] = 'DIRECCION'
        ws['H1'] = 'MUNICIPIO_CODIGO_DANE'
        ws['I1'] = 'TELEFONO'
        ws['J1'] = 'CELULAR'
        ws['K1'] = 'URL'
        ws['L1'] = 'GENERO'
        ws['M1'] = 'ESTRATO'
        ws['N1'] = 'FECHA_NACIMIENTO'
        ws['O1'] = 'MUNICIPIO_CODIGO_DANE_NACIMIENTO'
        ws['P1'] = 'TIPO_PERSONA'
        ws['Q1'] = 'ESTADO_PERSONA'
        ws['R1'] = 'OBSERVACIONES'

        total = len(personas)
        inicio = 2

        for persona in personas:
            ws.cell(row=inicio,column=1).value = persona.persona.tipoDocumento.tipo
            ws.cell(row=inicio,column=2).value = persona.persona.numeroDocumento
            ws.cell(row=inicio,column=3).value = persona.persona.nombres
            ws.cell(row=inicio,column=4).value = persona.persona.apellidos
            ws.cell(row=inicio,column=5).value = persona.persona.email
            ws.cell(row=inicio,column=6).value = ""
            ws.cell(row=inicio,column=7).value = persona.persona.direccion
            ws.cell(row=inicio,column=8).value = persona.persona.ciudadResidencia.codigo
            ws.cell(row=inicio,column=10).value = ""
            ws.cell(row=inicio,column=9).value = persona.persona.telefono
            ws.cell(row=inicio,column=11).value = ""
            ws.cell(row=inicio,column=12).value = persona.persona.sexo
            ws.cell(row=inicio,column=13).value = persona.persona.estrato
            ws.cell(row=inicio,column=14).value = persona.persona.fechaNacimiento
            ws.cell(row=inicio,column=15).value = persona.persona.ciudadNacimiento.codigo
            ws.cell(row=inicio,column=16).value = ""
            ws.cell(row=inicio,column=17).value = ""
            ws.cell(row=inicio,column=18).value = ""

            inicio = inicio + 1
        wb.save("/tmp/Reporte Cargue de datos.xlsx")

        fs = FileSystemStorage('/tmp')
        with fs.open('Reporte Cargue de datos.xlsx') as xls:
            response = HttpResponse(xls, content_type='application/ms-excel')
            response['Content-Disposition'] = 'filename="Reporte Cargue de datos.xlsx"'
            return response
    return render(request,'informes/report_cargue.html')
